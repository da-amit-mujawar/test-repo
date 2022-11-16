import pandas as pd 
from openpyxl import load_workbook
from openpyxl.styles import Font, colors, Alignment, Border, Side


# Auto-adjust columns' width
def set_column_width(df, sheet_name, excelwriter):
    for column in df:
        column_width = max(df[column].astype(str).map(len).max(), len(column))
        col_idx = df.columns.get_loc(column)
        excelwriter.sheets[sheet_name].set_column(col_idx, col_idx, column_width)


#Add thousand seperator
def add_thousand_separator(df):
    for col in df.columns:
        if df[col].dtype != 'object':
            df[col] = df[col].map('{:,}'.format)

    return df


#apply auto filter on top row of excel file 
def autoFilter(fileName):
    wb = load_workbook(fileName) 
    sheet = wb.active 
    #loop through all columns and rows to get max column and row number 
    for i in range (1,sheet.max_column+1):
        if i == 1:#if it is first column then only apply autofilter on that column otherwise ignore it 
            sheet.auto_filter.ref = "A1:" + chr(i+64) + str(sheet.max_row)  
        else:                              
            continue
